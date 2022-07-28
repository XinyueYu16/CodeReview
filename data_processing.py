# Read excel cells, including cell value, color, etc.
import openpyxl
import pandas as pd

workbook = openpyxl.load_workbook('...xlsx')
ttl_sheets = [sheetname for sheetname in workbook.sheetnames
              if sheetname.startswith('Total') or sheetname.find('GDP') != -1]
cur = workbook[ttl_sheets[0]]
rows, cols = cur.max_row, cur.max_column

whole = []
for i in range(5, 18+1):
    test = []
    for j in range(1, 21):
        #rows and cols start with 1 in Excel
        ce = cur.cell(row=i, column=j)
        fill = ce.fill
        if fill.start_color.rgb == "FFFFFF00":
            test.append(None)
        else:
            test.append(ce.value)
    whole.append(test)

df = pd.DataFrame(whole)
df.columns = df.loc[0, :]
df = df.loc[1:, :].copy().reset_index(drop=True)

# Check the missing data, plot their relationship
# Helpful when investigating different types of data missing 
import missingno as msnum

msnum.bar(city_data)
msnum.matrix(city_data)
msnum.heatmap(city_data)
